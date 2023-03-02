#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 18 09:09:39 2021

@author: johannesthellmann
"""
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

data = "/Users/johannesthellmann/Desktop/SoSe2021/Seminar BWL/data/VIX_Futures_month.xlsx"
wb = load_workbook(filename=data)

sheets = wb.sheetnames


#month names
#month = ['Jan', 'Feb','Mar','Apr','Mai','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
month = ['01','02','03', '04', '05', '06', '07', '08', '09', '10', '11', '12']


#available years
year = ['05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21']

#year = ['05', '06']


#VIX Future Zeitreihe
vix_f = pd.DataFrame()

#initial values 
last_day_month = pd.to_datetime("2004-12-01") #Hier anders, da davor keine Aufzeichnungen
month_counter = -1
#first col
firstcol =0
#lastcol
lastcol= 6 #range does not include the last one


for y in year:
    
    for sheet in sheets:
        month_counter = month_counter + 1
        if month_counter > 11:
            month_counter = 0
        
        ncol = [i for i in range(firstcol, lastcol)]
        #df_name = month[month_counter] + year[year_counter]
        df_name = pd.read_excel(data, sheet_name=sheet,header = 4, usecols=ncol)
        
        #rename columns since excel import gives different name when on same sheet
        col_names = df_name.columns
        df_name = df_name.rename(columns = {col_names[0]: 'Dates', 
                                  col_names[1]: 'CURRENT_CONTRACT_MONTH_YR', 
                                  col_names[2]: 'LAST_PRICE',
                                  col_names[3]: 'PX_BID',
                                  col_names[4]: 'PX_ASK',
                                  col_names[5]: 'PX_LAST'})
        #================for debug
        #print(df_name)
        
        #Neue Spalte für Maturity
        df_name['Maturity'] = df_name['Dates']
        
        #find last date: Maturity Date
        maturity_date=df_name['Dates'].max()
        
        #betreffende indices
        tage = df_name.index
        #Schreibe in jede Zelle mit index tage das Maturity Date
        df_name.loc[tage, ['Maturity']] = maturity_date
        
        #Verwende nur den Zeitabschnitt für die 1 month Future
        #Startday is last day of previous month
        start_date = last_day_month
        #endday ist letzter Tag diesen Monats
        

        
        #end_date = pd.to_datetime('20'+ year[y] + '-' + month[month_counter + 1]+ '-01') #zum testen
        #Nehme immer den ersten Tag des jeweiligen Monats
        end = ('20'+ y + '-' + month[month_counter]+ '-01')
        end_date = pd.to_datetime(end)
        
        #================for debug
        # print('startdate vor cut: ' + str(start_date))
        # print('enddate vor cut: '  + str(end_date))

        #Nur diesen Monat + letzter Tag des vorigen Monats
        mask1 = (df_name['Dates'] >= start_date) & (df_name['Dates'] < end_date)
        df_name = df_name.loc[mask1]
        
        #================for debug
        #print(df_name)
        
        #Finde letzten Tag diesen Monats und speichere für nächsten Durchlauf
        last_day_month = df_name['Dates'].max()
        
        #================for debug
        #print(str(last_day_month))
        
        #Entferne letzter Tag, da schon Future des nächsten Monats
        
        mask2 = (df_name['Dates'] >= start_date) & (df_name['Dates'] < last_day_month)
        df_name = df_name.loc[mask2]
        
        #Index erstellen aus Dates Spalte
        df_name['index'] = df_name['Dates']
        df_name = df_name.set_index(df_name['index'])
        df_name = df_name.drop(['index'], axis=1)
        
        #Hänge an die Zeitreihe der VIX Futures an
        vix_f = vix_f.append(df_name)
        
        print(sheet + ' ' + y)
        #================for debug
        # print('startdate: ' + str(start_date) + ', endate: ' + str(end_date))
        # print('lastdate month: ' + str(last_day_month))
                
    #Nächste Blattspalten
    if lastcol <= 118:
        firstcol = firstcol + 7
        lastcol = lastcol + 7
    else:
        break
        
        
vix_f.to_csv('/Users/johannesthellmann/Desktop/vix_f.csv', index=False)
        
        
        
        
        
        
        
        
        
        
        
    