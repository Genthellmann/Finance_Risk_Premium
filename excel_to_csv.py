#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 20 17:39:01 2021

@author: johannesthellmann
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

data = "/Users/johannesthellmann/Desktop/SoSe2021/Seminar BWL/data/VIX_data_MJ.xlsx"
wb = load_workbook(filename=data)

sheets = wb.sheetnames

vix_s = pd.read_excel(data, sheet_name='VIX',header = 5, usecols=[0,1])
vix_s.dropna(subset=["Dates"], inplace=True)

vix_s.to_csv('/Users/johannesthellmann/Desktop/vix_s.csv', index=False)
